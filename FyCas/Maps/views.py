import os
from . import models
from django.views.generic import *
from django.shortcuts import render
from django.urls import reverse_lazy, reverse
from django.shortcuts import redirect
from Customer.mixing import Options
from datetime import datetime
from django.template.loader import get_template
from django.http import HttpResponseRedirect, HttpResponse
from fpdf import FPDF
from django.http import JsonResponse
from twilio.rest import Client  
from Customer import models
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side,PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
# App Account
from oauth2client.service_account import ServiceAccountCredentials
from allauth.socialaccount.models import SocialAccount
from google.oauth2.credentials import Credentials

      
from google.oauth2 import service_account


from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build


# Suponiendo que ya tienes un objeto 'credentials' válido

class Maps(TemplateView, Options):
      template_name = "maps/information.html"

      
      
      def get(self, request, *args, **kwargs):
            nombre_archivo = "Info.docx"
            ruta_carpeta =  os.getcwd() + "\Clientes"
            # self.Send_WhatsApp_Message()
            return  self.Excel()
            # # Ejemplo de uso
            nombre_contacto = "Juan Perez"
            telefono_contacto = "+1234567890"
            # self.agregar_contacto(nombre_contacto, telefono_contacto)

            return super().get(request, *args, **kwargs)
      

      def Send_WhatsApp_Message(self):
            account_sid = 'AC32b5e94ce632aabd0a278a56e16bd44a'
            auth_token = 'a61c8b622e93ada5958d69dacef7c461'
            client = Client(account_sid, auth_token)

            Msm = "Recordario de Grupo Fycas \n \n Buenos días estimado(a), este es un recordatorio de su saldo pendiente \n \n Nombre: Xavier Torrero \n Cedula: 302-234433-0 \n Numero: (809)299-8306 \n \n Nos ponemos en contacto con usted para dar seguimiento al pago Numero dos (2) de veinte (20) cuotas. \n\n La fecha de vencimiento de su ultimo pago fue el 30/3/2024, a dia de hoy 6/4/2024 no hemos recibido el pago correspondiente. \n \n Para realizar el pago de la cuota pendiente mas mora actualmente es de; RD$5,500.00, puede llamar o dejar un mensaje via WhatsApp al +1 (809)870-7852. \n Atentamente: \n (Grupo Fycas)"
            message = client.messages.create(
                  body= Msm,
                  from_= '+13344384583',
                  to= '+18295577196' )
            
            # msg = client.messages.create(
            #       from_='whatsapp:+18295577196',
            #       body='Mensaje enviado por Wandy Olivares',
            #       to='whatsapp:+18295577196')
            return True



      def Excel(self):
            # Consultar el modelo de Django para obtener los datos.
            data = models.Customer.objects.all()
            date = datetime.now()
            
            # Crear un nuevo archivo Excel.
            workbook = Workbook()

            # Escribir los datos en el archivo Excel.
            for nombre_hoja in workbook.sheetnames:
                  worksheet = workbook[nombre_hoja]
                  bold_font = Font(bold=True)
                  
                  # Ajustar el ancho de las columnas
                  columnas = ["A", "B", "C", "D", "E", "F", "G"]
                  for columna in columnas:
                        worksheet.column_dimensions[columna].width = 20
                  worksheet.column_dimensions["A"].width = 25
                  worksheet.column_dimensions["A"].font = bold_font
                  
                  worksheet.column_dimensions["B"].width = 30
                  worksheet.column_dimensions["C"].width = 20
                  worksheet.column_dimensions["D"].width = 20
                  worksheet.column_dimensions["E"].width = 10
                  worksheet.column_dimensions["F"].width = 20
                  worksheet.column_dimensions["G"].width = 20
                  # Aplicar formato de fuente a la fila 3


                  worksheet.append(["TIPO DE ENTIDAD", "NOMBRE DEL CLIENTE", "APELLIDOS", "CEDULA O RNC", "SEXO", "ESTADO CIVIL", "OCUPACION", "CODIGO DE CLIENTE", "FECHA DE NACIMIENTO", "NACIONALIDAD", "DIRECCION", "SECTOR", "CALLE/NUMERO", "MUNICIPIO", "CIUDAD", "PROVINCIA", "PAIS", "DIR_Referencia", "TELEFONO", 'EMPRESA DONDE TRABAJA', "CARGO", "DIRECCION", "SECTOR", "CALLE/NUMERO", "MUNICIPIO", "CIUDAD", "PROVINCIA", "PAIS"])

                  # Insertar una nueva fila en la fila 1
                  worksheet.insert_rows(1)
                  # Combinar las celdas de la fila 1
                  worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=15)
         
                  # Insertar una nueva fila en la fila 2
                  worksheet.insert_rows(2)
                  worksheet.insert_rows(3)
                  worksheet.insert_rows(4)
                  worksheet.insert_rows(5)
                  # Combinar las celdas de la fila 2
                  worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=15)

                  worksheet.cell(row=1, column=1).value = "GRUPO FYCAS, SRL"
                  worksheet.cell(row=1, column=1).font = Font(name="Arial", size=18, bold=True, color="000000")

                  worksheet.cell(row=2, column=1).value = "AF000024759"
                  worksheet.cell(row=2, column=1).font = Font(name="Arial", size=14)
                            
                  for cell in worksheet[1]:
                     cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                  for cell in worksheet[2]:
                     cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                  for cell in worksheet[3]:
                     cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                  for cell in worksheet[4]:
                     cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                  for cell in worksheet[5]:
                     cell.fill = PatternFill(start_color="0080ff", end_color="0080ff", fill_type="solid")
   
                  
                  worksheet.cell(row=5, column=1).value = "Datos Personales"
                  worksheet.cell(row=5, column=1).font =  Font(name="Cambria", size=16, bold=True, color="FFFFFF")
                  
         
                  for col in range(1, 8):  # Iterate from column B to G (1-based indexing)
                        cell = worksheet.cell(row=6, column=col)
                        cell.font = Font(name="Calibri", size=14,  bold=True)
                        
                  for row in data:
                        worksheet.append([row.type_input, row.name, row.last_name, row.dni, row.sexo, row.estado_civil, row.ocupacion])
                        
                        

            # Enviar el archivo Excel al usuario como respuesta HTTP.
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=Grupo Fycas SRL {date}.xlsx'
            workbook.save(response)
            return response
      
      

# Define los alcances que necesitas para acceder a los contactos



      def obtener_credenciales(self):
            SCOPES = ['https://www.googleapis.com/auth/contacts']
            
            # Intenta cargar las credenciales desde el archivo token.json
            creds = None
            if os.path.exists('credentials.json'):
                  creds = Credentials.from_authorized_user_file('credentials.json')

            # Si no hay credenciales disponibles o no son válidas, inicia el flujo de autorización
            if not creds or not creds.valid:
                  if creds and creds.expired and creds.refresh_token:
                        creds.refresh(Request())
                  else:
                        # Si no hay credenciales disponibles, inicia el flujo de autorización
                        flow = InstalledAppFlow.from_client_secrets_file('Customer/Credentials-Apis/credentials.json', SCOPES, redirect_uri='http://localhost:56083')
                        
                        creds = flow.run_local_server(port=56083)

                  # Guarda las credenciales en credentials.json para futuros usos
                  with open('credentials.json', 'w') as token:
                        token.write(creds.to_json())

            return creds

      def agregar_contacto(self, nombre, telefono):
            creds = self.obtener_credenciales()
            # Crea un objeto de servicio para interactuar con la API de Google Contacts
            service = build('people', 'v1', credentials=creds)
            # Define el contacto
            contacto = {
                  "names": [
                        {
                        "givenName": nombre
                        }
                  ],
                  "phoneNumbers": [
                        {
                        "value": telefono,
                        "type": "mobile"
                        }
                  ]
            }
            # Agrega el contacto
            service.people().createContact(body=contacto).execute()
            print("Contacto agregado exitosamente.")
